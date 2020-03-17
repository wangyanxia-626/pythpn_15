# -*- coding:utf-8 -*-
# @Time :2020-03-12 16:16
# @Email :876417305@qq.com
# @Author :yanxia
# @File :do_excel.PY
#完成excel的读和写
import openpyxl
from class_0312 import http_request
class Case:
    """测试用例类，每个测试用例，实际上就是它的一个实例（每一个测试用例，实际就是Case的一个实例）"""
    def __init__(self):
        self.case_id=None
        self.title=None
        self.url=None
        self.data=None
        self.method=None
        self.expected=None
        self.actual=None
        self.result=None

class DoExcel:
    def __init__(self,file_name,sheet_name):
        try:
            self.file_name=file_name
            self.workbook=openpyxl.load_workbook(file_name)
            self.sheet_name=sheet_name
            self.sheet=self.workbook[sheet_name]
        except Exception as e:
            print("请核对excel路径")
            raise e
    def get_cases(self):
        max_row=self.sheet.max_row #获取最大行数
        cases=[] #列表存放所有的测试用例
        for i in range(2,max_row+1):
            case=Case()#Case实例化
            case.case_id=self.sheet.cell(row=i,column=1).value
            case.title = self.sheet.cell(row=i, column=2).value
            case.url = self.sheet.cell(row=i, column=3).value
            case.data= self.sheet.cell(row=i, column=4).value
            case.method= self.sheet.cell(row=i, column=5).value
            case.expected= self.sheet.cell(row=i, column=6).value
            cases.append(case)
        self.workbook.close()
        return cases
    def write_result(self,row,actual,result):#列固定，行不固定
        sheet=self.workbook[self.sheet_name]#因为写完保存之后要关闭这个excel
        sheet.cell(row,7).value=actual
        sheet.cell(row,8).value=result
        self.workbook.save(filename=self.file_name)
        self.workbook.close()

if __name__ == '__main__':
    do_excel=DoExcel('cases.xlsx',sheet_name='login')
    cases=do_excel.get_cases() #实际是一个对象实例，看整个数据可以用case.__dict_
    http_request=http_request.HttpRequest()# _
    for case in cases:
        # print(case.case_id)
        # print(case.method)
        # print(case.data)
        print(case.__dict__)#每一个类都有这个dict,会返回类里面所有属性
        resp=http_request.request(case.method,case.url,case.data)
        actual=resp.text
        if case.expected==actual:#判断期望结果与实际结果一致
            do_excel.write_result(case.case_id+1,actual,"pass")
        else:
            do_excel.write_result(case.case_id+1,actual,"fail")

