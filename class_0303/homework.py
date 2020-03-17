# -*- coding:utf-8 -*-
# @Time :2020-03-04 18:12
# @Email :876417305@qq.com
# @Author :yanxia
# @File :homework.PY
#安排一个作业 #写一个类 类里面有2个方法 1）读数据 2）写数据
#1）读数据可以读取整个Excel里面所有的数据，每一行数据都放到一个子列表里面，
# 所有子列表数据放到一个大列表里面,要求把读取到的数据返回
# #2）写数据可以在Excel里面指定的单元格里面写入指定的值，不需要返回值
#温馨提示：记得关闭和保存Excel

from openpyxl import load_workbook
# class ReadExcel:
#     def __init__(self,wb,sheet):
#         self.wb=wb
#         self.sheet=sheet
#
#     def read_excel(self):
#         max_list=[]
#         for i in range(1,self.sheet.max_row+1):
#             sub_list=[]
#             for j in range(1,self.sheet.max_column+1):
#                 #if self.sheet.cell(i.j).value:
#                     sub_list.append(self.sheet.cell(i, j).value)
#             max_list.append(sub_list)
#         return max_list
#
#     def write_excel(self):
#         self.sheet.cell(6,1,"这个插入值")
#         self.wb.save()
#         self.wb.close()
# if __name__ == '__main__':
#     wb=load_workbook("py15.xlsx")
#     sheet=wb["Sheet1"]
#     print(ReadExcel(wb,sheet).read_excel())


#老师讲解
class DoExcel:
    def __init__(self,file_name,sheet_name):
        self.file_name=file_name
        self.sheet_name=sheet_name
    def read_data(self):
        """
        :param file_name:目标工作薄的名称
        :param sheet_name:指定的表单名称
        :return:返回列表形式的测试数据
        """
        wb=load_workbook(self.file_name)
        sheet=wb[self.sheet_name]
        all_data=[]
        for i in range(1,sheet.max_row+1):
            sub_list = []
            sub_list.append(sheet.cell(i,1).value)
            sub_list.append(sheet.cell(i,2).value)
            sub_list.append(sheet.cell(i,3).value)
            sub_list.append(sheet.cell(i,4).value)
            all_data.append(sub_list)
        wb.close()
        return all_data
    def write_back(self,row,column,new_value):#写回数据的方法
        wb=load_workbook(self.file_name)
        sheet=wb[self.sheet_name]
        #指定的行列值写入指定的值
        sheet.cell(row,column).value=new_value
        wb.save(self.file_name)
        wb.close()

if __name__ == '__main__':
    t=DoExcel("py15.xlsx","Sheet1")
    t.write_back(1,2,"python")
    all_data=t.read_data()
    print(all_data)




























