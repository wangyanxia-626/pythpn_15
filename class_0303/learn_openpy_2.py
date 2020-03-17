# -*- coding:utf-8 -*-
# @Time :2020-03-04 17:27
# @Email :876417305@qq.com
# @Author :yanxia
# @File :learn_openpy_2.PY
from openpyxl import load_workbook
wb=load_workbook("py15.xlsx")
sheet=wb["Sheet1"]
# 循环取值第二种办法
print(sheet.max_row)#获取最大行
print(sheet.max_column)#获取最大列
# 循环取值第一种方法
for i in range(1,sheet.max_row+1):
        for j in range(1,sheet.max_column+1):
                if sheet.cell(i,j).value: # if条件为非空，为真的时候才执行if后面的语句
                        res=(sheet.cell(i,j).value)
                        print(res)

