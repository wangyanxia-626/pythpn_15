# -*- coding:utf-8 -*-
# @Time :2020-03-04 10:27
# @Email :876417305@qq.com
# @Author :yanxia
# @File :learn_openpyxl.PY
# 创建excel文件的模块workbook
#读写excel的文件模块load_workbook
# from openpyxl import workbook
# # 创建excel文件
# wb=workbook.Workbook()
# wb.create_sheet("lbb")# 创建表单的方法
# wb.save("py15_lbb.xlsx")# 另存为
# 开始读写的操作
from openpyxl import load_workbook
#读操作 三步走
#第一步 打开excel工作薄---workbook
wb=load_workbook("py15.xlsx")
# 第二步 定位到表单
sheet=wb["Sheet1"]
# 第三步 定位单元格 获取内容 根据行列坐标获取值
res=sheet.cell(1,2).value #获取的第1行第2列
res1=sheet.cell(3,1).value
print("res的值是：{},res的类型是：{}".format(res1,type(res1)))
#得出结论：数字还是数字 其他数据类型全是字符串类型
res2=eval(sheet.cell(4,3).value)
print(res2,type(res2))
# eval()可以把数据转成python原本可以识别的数据类型，但要注意普通字符串与字符串的区别
# 写入值 修改和新增都是这个方法 操作完成之后一定要关闭
sheet.cell(5,1).value="白日依山尽"
sheet.cell(5,2,"yanxia")
#保存%另存为 如果是保存到当前的excel的话，记得要关闭excel 不然会报错permission
wb.save("py15.xlsx")
# 操作完成之后关闭文件
wb.close()
#循环读值
