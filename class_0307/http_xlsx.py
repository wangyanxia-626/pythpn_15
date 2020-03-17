# -*- coding:utf-8 -*-
# @Time :2020-03-09 11:08
# @Email :876417305@qq.com
# @Author :yanxia
# @File :http_xlsx.PY
from openpyxl import load_workbook
import json

def get_datavalue():
    wb=load_workbook("data.xlsx")
    #sheet=wb["Sheet1"]
    sheet=wb.worksheets[0]#是一个列表，取第一个
    login_data=[]
    for row in range(2,sheet.max_row+1):
        user_login={"url":sheet.cell(row,1).value,
                    "method":sheet.cell(row,2).value,
                    "data":json.loads(sheet.cell(row,3).value),
                    "expected":sheet.cell(row,4).value,
                    "case_id":sheet.cell(row,7).value}
        login_data.append(user_login)
        wb.close()
        return login_data
def write_result(row,result):
    wb = load_workbook("data.xlsx")
    # sheet=wb["Sheet1"]
    sheet = wb.worksheets[0]  # 是一个列表，取第一个
    sheet.cell(row,6,result)
    wb.save("data.xlsx")
    wb.close()