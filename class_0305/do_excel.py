# -*- coding:utf-8 -*-
# @Time :2020-03-05 17:46
# @Email :876417305@qq.com
# @Author :yanxia
# @File :do_excel.PY
from openpyxl import load_workbook
from class_0305.read_config import ReadConfig
class DoExcel:
    def __init__(self,file_name,sheet_name):
        self.file_name=file_name
        self.sheet_name=sheet_name
    def read_data(self):
        """
        :param file_name:目标工作薄的名称
        :param sheet_name:指定的表单名称
        :return:返回列表形式的测试数据
        """
        wb=load_workbook(self.file_name)
        sheet=wb[self.sheet_name]
        #从配置文件读取数据 决定获取哪些数据
        line = ReadConfig("case.conf").get_strValue("lineno", "line")
        all_data=[]
        for i in range(1,sheet.max_row+1):
            sub_list = []
            sub_list.append(sheet.cell(i,1).value)
            sub_list.append(sheet.cell(i,2).value)
            sub_list.append(sheet.cell(i,3).value)
            sub_list.append(sheet.cell(i,4).value)
            all_data.append(sub_list)
        final_data=[]
        if line=="all":
            final_data=all_data
        else:#读取配置文件里指定列表里面的指定行数的数据
            for i in eval(line):#遍历列表里面行数的数字，是把字符串的line，变成列表类型的line
                final_data.append(all_data[i-1])
                #添加数据，line列表里面的数字 跟all_data的数据的索引是i-1的关系（其实就是all_data里面第一行对应索引的0）
        wb.close()
        return final_data
    def write_back(self,row,column,new_value):#写回数据的方法
        wb=load_workbook(self.file_name)
        sheet=wb[self.sheet_name]
        #指定的行列值写入指定的值
        sheet.cell(row,column).value=new_value
        wb.save(self.file_name)
        wb.close()

if __name__ == '__main__':
    t=DoExcel("py15.xlsx","Sheet1").read_data()
    # t.write_back(1,2,"python")
    print(t)
