# -*- coding:utf-8 -*-
# @Time :2020-03-13 11:58
# @Email :876417305@qq.com
# @Author :yanxia
# @File :do_excel.PY
import openpyxl
from class_0314.common import http_request
from class_0314.common import contants


class Case:
    def __init__(self):
        self.case_id = None
        self.title = None
        self.url = None
        self.data = None
        self.method = None
        self.expected = None
        self.actual = None
        self.result = None
        self.sql=None


class DoExcel:
    def __init__(self, file_name, sheet_name):
        self.file_name = file_name
        self.workbook = openpyxl.load_workbook(file_name)
        self.sheet_name = sheet_name
        self.sheet = self.workbook[sheet_name]

    def get_cases(self):
        max_row = self.sheet.max_row
        cases = []
        for i in range(2, max_row + 1):
            case = Case()
            case.case_id = self.sheet.cell(row=i, column=1).value
            case.title = self.sheet.cell(row=i, column=2).value
            case.url = self.sheet.cell(row=i, column=3).value
            case.data = self.sheet.cell(row=i, column=4).value
            case.method = self.sheet.cell(row=i, column=5).value
            case.expected = self.sheet.cell(row=i, column=6).value
            case.sql=self.sheet.cell(row=i,column=9).value#执行的sql
            cases.append(case)
        self.workbook.close()
        return cases

    def write_result(self, row, actual, result):
        sheet = self.workbook[self.sheet_name]
        sheet.cell(row, 7).value = actual
        sheet.cell(row, 8).value = result
        self.workbook.save(filename=self.file_name)
        self.workbook.close()


if __name__ == '__main__':
    do_excel = DoExcel(contants.case_file, sheet_name="recharge")
    cases = do_excel.get_cases()
    http_request = http_request.HttpRequest()
    for case in cases:
        print(case.__dict__)
        resp = http_request.request(case.method, case.url, case.data)
        actual = resp.text
        if case.expected == actual:  # 判断期望结果与实际结果一致
            do_excel.write_result(case.case_id + 1, actual, "PASS")
        else:
            do_excel.write_result(case.case_id + 1, actual, "FAIL")
